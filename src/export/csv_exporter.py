"""
Export module — CSV, Excel (.xlsx) and PDF for job postings and prospects.
Auto-export to CSV is triggered after each pipeline run.
Manual export is available from the web dashboard.
"""

import csv
import io
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from src.database import db

logger = logging.getLogger(__name__)

EXPORTS_DIR = Path(__file__).parent.parent.parent / "exports"

# Column definitions
PROSPECT_COLUMNS = [
    "date",
    "company_name",
    "company_domain",
    "contact_name",
    "email",
    "title",
    "job_posting_title",
    "keyword",
    "email_status",
    "outreach_status",
]

PROSPECT_HEADERS = [
    "Date",
    "Company",
    "Domain",
    "Contact Name",
    "Email",
    "Title / Position",
    "Job Posting",
    "Keyword",
    "Email Status",
    "Outreach Status",
]

POSTING_COLUMNS = [
    "source",
    "external_id",
    "title",
    "company_name",
    "company_domain",
    "location",
    "keyword_matched",
    "published_at",
    "scraped_at",
    "url",
]

POSTING_HEADERS = [
    "Source",
    "External ID",
    "Job Title",
    "Company",
    "Domain",
    "Location",
    "Keyword",
    "Published",
    "Scraped",
    "URL",
]


def _ensure_exports_dir() -> Path:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    return EXPORTS_DIR


def _ts() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


# ------------------------------------------------------------------
# CSV
# ------------------------------------------------------------------

def export_prospects_csv(filename: str = None) -> Optional[str]:
    """Export all prospects to CSV file. Returns path or None."""
    try:
        out_dir = _ensure_exports_dir()
        filepath = out_dir / (filename or f"prospects_{_ts()}.csv")
        rows = db.get_prospects_for_export()
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=PROSPECT_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow(row)
        logger.info("CSV: exported %d prospects to %s", len(rows), filepath)
        return str(filepath)
    except Exception as exc:
        logger.error("CSV export (prospects) failed: %s", exc)
        return None


def export_postings_csv(filename: str = None) -> Optional[str]:
    """Export all job postings to CSV file. Returns path or None."""
    try:
        out_dir = _ensure_exports_dir()
        filepath = out_dir / (filename or f"postings_{_ts()}.csv")
        rows = db.get_postings_for_export()
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=POSTING_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow(row)
        logger.info("CSV: exported %d postings to %s", len(rows), filepath)
        return str(filepath)
    except Exception as exc:
        logger.error("CSV export (postings) failed: %s", exc)
        return None


def stream_prospects_csv():
    """Yield CSV bytes for streaming HTTP response."""
    rows = db.get_prospects_for_export()
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=PROSPECT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    yield buf.getvalue().encode("utf-8-sig")
    for row in rows:
        buf.seek(0); buf.truncate()
        writer.writerow(row)
        yield buf.getvalue().encode("utf-8")


def stream_postings_csv():
    """Yield CSV bytes for streaming HTTP response."""
    rows = db.get_postings_for_export()
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=POSTING_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    yield buf.getvalue().encode("utf-8-sig")
    for row in rows:
        buf.seek(0); buf.truncate()
        writer.writerow(row)
        yield buf.getvalue().encode("utf-8")


# ------------------------------------------------------------------
# Excel (.xlsx)
# ------------------------------------------------------------------

def _style_xlsx_header(ws, columns: list[str]):
    """Apply header styling to the first row of an openpyxl worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(bottom=thin)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = border


def _style_xlsx_rows(ws, num_cols: int, num_rows: int):
    """Apply alternating row colours and borders."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    even_fill = PatternFill("solid", fgColor="F0F4F8")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")
    normal_font = Font(size=9)
    thin = Side(style="thin", color="E2E8F0")
    border = Border(bottom=thin)

    for row_idx in range(2, num_rows + 2):
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.font = normal_font
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = border


def build_prospects_xlsx() -> bytes:
    """Build and return Excel workbook bytes for all prospects."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    rows = db.get_prospects_for_export()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospects"
    ws.freeze_panes = "A2"

    _style_xlsx_header(ws, PROSPECT_HEADERS)

    for row in rows:
        ws.append([row.get(col, "") or "" for col in PROSPECT_COLUMNS])

    _style_xlsx_rows(ws, len(PROSPECT_COLUMNS), len(rows))

    # Column widths
    widths = [18, 28, 26, 24, 34, 22, 38, 14, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_postings_xlsx() -> bytes:
    """Build and return Excel workbook bytes for all job postings."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    rows = db.get_postings_for_export()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Postings"
    ws.freeze_panes = "A2"

    _style_xlsx_header(ws, POSTING_HEADERS)

    for row in rows:
        ws.append([row.get(col, "") or "" for col in POSTING_COLUMNS])

    _style_xlsx_rows(ws, len(POSTING_COLUMNS), len(rows))

    widths = [10, 14, 38, 28, 26, 20, 14, 20, 20, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------------
# PDF
# ------------------------------------------------------------------

def _pdf_table(data: list[list], col_widths: list, title: str, subtitle: str) -> bytes:
    """
    Build a styled PDF with a title and a data table.
    Returns PDF bytes.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_LEFT

    buf = io.BytesIO()
    page = landscape(A4)
    doc = SimpleDocTemplate(
        buf,
        pagesize=page,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Normal"],
        fontSize=14,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#1E3A5F"),
        spaceAfter=2,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.HexColor("#64748B"),
        spaceAfter=6,
    )
    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
        alignment=TA_LEFT,
    )

    # Wrap cell text in Paragraphs so long values wrap properly
    header_row = data[0]
    body_rows  = data[1:]

    table_data = [header_row]
    for row in body_rows:
        table_data.append([
            Paragraph(str(cell) if cell else "", cell_style)
            for cell in row
        ])

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        # Header
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 7),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING",    (0, 0), (-1, 0), 6),
        # Body
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 1), (-1, -1), 7),
        ("TOPPADDING",    (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
        # Grid
        ("GRID",        (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("VALIGN",      (0, 0), (-1, -1), "TOP"),
    ]))

    story = [
        Paragraph(title, title_style),
        Paragraph(subtitle, subtitle_style),
        Spacer(1, 4 * mm),
        tbl,
    ]
    doc.build(story)
    return buf.getvalue()


def build_prospects_pdf() -> bytes:
    """Build and return PDF bytes for all prospects."""
    from reportlab.lib.units import mm

    rows = db.get_prospects_for_export()
    ts = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M")

    data = [PROSPECT_HEADERS]
    for row in rows:
        data.append([row.get(col, "") or "" for col in PROSPECT_COLUMNS])

    # Column widths in mm (landscape A4 = ~277mm usable)
    col_widths = [22*mm, 34*mm, 28*mm, 28*mm, 38*mm, 26*mm, 42*mm, 16*mm, 18*mm, 20*mm]

    return _pdf_table(
        data,
        col_widths,
        title="Sperton Leads — Prospects",
        subtitle=f"Exported {ts}  •  {len(rows)} records  •  sperton.com",
    )


def build_postings_pdf() -> bytes:
    """Build and return PDF bytes for all job postings."""
    from reportlab.lib.units import mm

    rows = db.get_postings_for_export()
    ts = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M")

    data = [POSTING_HEADERS]
    for row in rows:
        data.append([row.get(col, "") or "" for col in POSTING_COLUMNS])

    # Column widths in mm (landscape A4 = ~277mm usable)
    col_widths = [12*mm, 16*mm, 42*mm, 30*mm, 24*mm, 20*mm, 16*mm, 20*mm, 20*mm, 50*mm]

    return _pdf_table(
        data,
        col_widths,
        title="Sperton Leads — Job Postings",
        subtitle=f"Exported {ts}  •  {len(rows)} records  •  sperton.com",
    )


# ------------------------------------------------------------------
# Auto-export (pipeline hook)
# ------------------------------------------------------------------

def auto_export_after_run() -> Optional[str]:
    """Called at the end of a pipeline run. Saves CSV + XLSX."""
    csv_path = export_prospects_csv()
    try:
        out_dir = _ensure_exports_dir()
        ts = _ts()
        xlsx_path = out_dir / f"prospects_{ts}.xlsx"
        xlsx_path.write_bytes(build_prospects_xlsx())
        logger.info("Auto-export XLSX saved to %s", xlsx_path)
    except Exception as exc:
        logger.warning("Auto-export XLSX failed: %s", exc)
    return csv_path


# Backward compat aliases
export_prospects = export_prospects_csv


# ------------------------------------------------------------------
# ERA Group PDF Extractions Export
# ------------------------------------------------------------------

ERA_EXTRACTION_COLUMNS = [
    "filename",
    "extraction_type",
    "confidence_score",
    "extraction_date",
    "invoice_number",
    "invoice_date",
    "due_date",
    "vendor_name",
    "buyer_name",
    "currency",
    "subtotal",
    "tax_amount",
    "total_amount",
    "payment_terms",
    "line_item_count",
    "pages_processed",
]

ERA_EXTRACTION_HEADERS = [
    "File",
    "Doc Type",
    "Confidence",
    "Extraction Date",
    "Invoice #",
    "Invoice Date",
    "Due Date",
    "Vendor",
    "Buyer",
    "Currency",
    "Subtotal",
    "Tax",
    "Total",
    "Payment Terms",
    "Line Items",
    "Pages",
]


def _parse_extraction_data(extraction: dict) -> dict:
    """Parse extracted_data JSON and flatten to a row."""
    import json as _json

    raw = extraction.get("extracted_data", "{}")
    if isinstance(raw, str):
        try:
            data = _json.loads(raw)
        except Exception:
            data = {}
    else:
        data = raw or {}

    row = {col: extraction.get(col, "") for col in ERA_EXTRACTION_COLUMNS}
    # Overlay extracted fields
    for key in ["invoice_number", "invoice_date", "due_date", "vendor_name", "buyer_name",
                 "currency", "subtotal", "tax_amount", "total_amount", "payment_terms"]:
        if data.get(key):
            row[key] = data[key]

    # Derived counts
    row["line_item_count"] = len(data.get("line_items", []))
    row["pages_processed"] = data.get("pages_processed", extraction.get("page_number", 1))

    # Confidence as percentage
    conf = extraction.get("confidence_score")
    if conf is not None:
        row["confidence_score"] = f"{float(conf) * 100:.0f}%"

    return row


def export_era_extractions_csv(extractions: list[dict] = None) -> io.StringIO:
    """
    Export ERA Group extractions to CSV with flattened extracted data fields.
    """
    if extractions is None:
        extractions = db.get_all_extractions_for_export()

    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=ERA_EXTRACTION_COLUMNS,
        extrasaction="ignore",
    )

    writer.writerow(dict(zip(ERA_EXTRACTION_COLUMNS, ERA_EXTRACTION_HEADERS)))

    for extraction in extractions:
        row = _parse_extraction_data(extraction)
        writer.writerow(row)

    logger.info(f"Exported {len(extractions)} ERA extractions to CSV")
    return output


def build_era_extractions_xlsx(extractions: list[dict] = None) -> bytes:
    """
    Build styled Excel workbook for ERA extractions with extracted data fields.
    Includes a summary sheet + a line items sheet.
    """
    if extractions is None:
        extractions = db.get_all_extractions_for_export()

    import json as _json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Extractions"

    header_fill = PatternFill(start_color="00537F", end_color="00537F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_fill_light = PatternFill(start_color="F0F6FF", end_color="F0F6FF", fill_type="solid")
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # Headers
    for col_idx, header in enumerate(ERA_EXTRACTION_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    for row_idx, extraction in enumerate(extractions, 2):
        row = _parse_extraction_data(extraction)
        for col_idx, column in enumerate(ERA_EXTRACTION_COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row.get(column, "")
            if row_idx % 2 == 0:
                cell.fill = data_fill_light
            cell.alignment = data_alignment
            cell.border = border

    ws.freeze_panes = "A2"
    widths = [30, 12, 10, 14, 14, 14, 14, 24, 24, 8, 12, 12, 14, 20, 8, 6]
    for i, w in enumerate(widths, start=1):
        if i <= len(ERA_EXTRACTION_COLUMNS):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Line Items ──
    ws2 = wb.create_sheet("Line Items")
    li_headers = ["Source File", "Line #", "Product Code", "Description", "Qty", "Unit", "Unit Price", "Tax", "Amount"]

    for col_idx, header in enumerate(li_headers, 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    li_row = 2
    for extraction in extractions:
        raw = extraction.get("extracted_data", "{}")
        if isinstance(raw, str):
            try:
                data = _json.loads(raw)
            except Exception:
                data = {}
        else:
            data = raw or {}

        for item in data.get("line_items", []):
            values = [
                extraction.get("filename", ""),
                item.get("line_no", ""),
                item.get("product_code", ""),
                item.get("description", ""),
                item.get("quantity", ""),
                item.get("unit", ""),
                item.get("unit_price", ""),
                item.get("tax", ""),
                item.get("amount", ""),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws2.cell(row=li_row, column=col_idx)
                cell.value = val
                if li_row % 2 == 0:
                    cell.fill = data_fill_light
                cell.alignment = data_alignment
                cell.border = border
            li_row += 1

    ws2.freeze_panes = "A2"
    li_widths = [30, 6, 14, 40, 8, 8, 12, 10, 14]
    for i, w in enumerate(li_widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info(f"Built ERA extractions XLSX with {len(extractions)} records")
    return output.getvalue()
export_postings  = export_postings_csv
